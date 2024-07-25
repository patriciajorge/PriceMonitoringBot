from selenium.common.exceptions import NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from time import sleep
import openpyxl
import schedule
from datetime import datetime
import re

def start_driver():
    chrome_options = Options()
    arguments = ['--lang=en-US', '--window-size=1300,1000', '--incognito']
    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,
    })
    driver = webdriver.Chrome(options=chrome_options)

    wait = WebDriverWait(
        driver,
        10,
        poll_frequency=1,
        ignored_exceptions=[NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException]
    )
    return driver, wait

def enter_site(driver, wait, workbook):
    print('Entering the Yamaha website')
    driver.get('https://www.yamaha-motor.com.br/mt-03-abs/product/30121')

    print('Typing the ZIP code')
    zip_field = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='x-input-localization']")))
    zip_field.send_keys('13880000')
    zip_field.send_keys(Keys.ENTER)
    sleep(4)
    # scroll down the page
    driver.execute_script("window.scrollTo(0,4900);")

    sleep(4)
    try:
        print('Checking the price')
        product_element = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='x-product-title']/span")))
        product = product_element[0].text

        price_element = wait.until(EC.visibility_of_element_located((By.XPATH, "//span[@class='x-product-price_value'][1]")))
        price_text = price_element.text

        # Extract numeric value from price text
        price = re.sub(r'[^\d,]', '', price_text).replace(',', '.')
        price = float(price)

        link = 'https://www.yamaha-motor.com.br/checkout'

        date = datetime.now().strftime('%d-%m-%Y')
        
    except NoSuchElementException:
        print("Price not found")
    
    save_data_excel(workbook, product, price, link, date)
    driver.close()

def save_data_excel(workbook, product, price, link, date):
    try:
        print('Saving the motorcycle information in the Excel sheet')
        product_page = workbook['Motorcycle Price']
    except KeyError:
        print('Creating the Excel sheet')
        product_page = workbook.create_sheet('Motorcycle Price')

    product_page['A1'].value = 'Motorcycle'
    product_page['B1'].value = 'Price'
    product_page['C1'].value = 'Link'
    product_page['D1'].value = 'Current_Date'

    product_page['A2'].value = product
    product_page['B2'].value = price
    product_page['C2'].value = link
    product_page['D2'].value = date

    workbook.save('motorcycle_price.xlsx')

def start_automation():
    # Load or create the workbook before starting automation
    try:
        workbook = openpyxl.load_workbook('motorcycle_price.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save('motorcycle_price.xlsx')
    
    driver, wait = start_driver()

    try:
        enter_site(driver, wait, workbook)
    finally:
        driver.quit()

if __name__ == "__main__":
    schedule.every(30).minutes.do(start_automation)

    while True:
        schedule.run_pending()
        sleep(1)
